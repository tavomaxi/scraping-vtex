#!/usr/bin/env python3
"""
Portsaid Catalog Scraper - Herramienta automatizada de extracción de catálogo
Autor: Asistente AI
Versión: 1.0.0
"""

import requests
import json
import time
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import DataBarRule
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from io import BytesIO


class PortsaidScraper:
    """Scraper automatizado para catálogo Portsaid."""
    
    def __init__(self, config_path: Optional[str] = None):
        """Inicializa el scraper con configuración."""
        self.base_url = "https://www.portsaid.com.ar"
        self.api_endpoint = "/api/io/_v/api/intelligent-search/product_search"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json'
        }
        
        # Configurar logging
        self.setup_logging()
        
        # Configuración por defecto
        self.config = {
            'output_dir': 'output',
            'sample_images_count': 30,
            'request_delay': 0.3,
            'timeout': 10,
            'max_retries': 3
        }
        
        # Cargar configuración personalizada si existe
        if config_path and os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                self.config.update(json.load(f))
        
        self.logger.info(f"Scraper inicializado. Config: {self.config}")
    
    def setup_logging(self):
        """Configura el sistema de logging."""
        log_dir = Path("logs")
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"scraper_{timestamp}.log"
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def fetch_products(self) -> List[Dict]:
        """Extrae todos los productos de la API."""
        all_products = []
        page = 1
        total_pages = None
        
        self.logger.info("Iniciando extracción de productos...")
        
        while True:
            url = f"{self.base_url}{self.api_endpoint}?page={page}"
            
            for attempt in range(self.config['max_retries']):
                try:
                    response = requests.get(
                        url, 
                        headers=self.headers, 
                        timeout=self.config['timeout']
                    )
                    response.raise_for_status()
                    data = response.json()
                    
                    products = data.get('products', [])
                    if not products:
                        self.logger.info("No hay más productos.")
                        break
                    
                    all_products.extend(products)
                    
                    if total_pages is None:
                        total_products = data.get('recordsFiltered', 0)
                        per_page = data.get('pagination', {}).get('perPage', 24)
                        total_pages = (total_products + per_page - 1) // per_page
                        self.logger.info(f"Total estimado: {total_products} productos en {total_pages} páginas")
                    
                    self.logger.info(f"Página {page}/{total_pages}: {len(products)} productos")
                    break
                    
                except requests.RequestException as e:
                    self.logger.warning(f"Intento {attempt + 1} fallido para página {page}: {e}")
                    if attempt == self.config['max_retries'] - 1:
                        self.logger.error(f"No se pudo obtener página {page} después de {self.config['max_retries']} intentos")
                        raise
                    time.sleep(2 ** attempt)  # Backoff exponencial
            
            if not products:
                break
            
            page += 1
            time.sleep(self.config['request_delay'])
        
        self.logger.info(f"Extracción completada: {len(all_products)} productos")
        return all_products
    
    def process_products(self, products: List[Dict]) -> List[Dict]:
        """Procesa los productos crudos en formato estructurado."""
        processed = []
        
        for product in products:
            try:
                # Precios
                price_range = product.get('priceRange', {})
                selling_price = price_range.get('sellingPrice', {}).get('highPrice', 0)
                list_price = price_range.get('listPrice', {}).get('highPrice', 0)
                
                # Categorías
                categories = product.get('categories', [])
                category = categories[0] if categories else ''
                
                # Imagen
                items = product.get('items', [])
                main_image = ''
                if items:
                    images = items[0].get('images', [])
                    if images:
                        main_image = images[0].get('imageUrl', '')
                
                # Talles
                sku_specs = product.get('skuSpecifications', [])
                sizes = []
                for spec in sku_specs:
                    field_name = spec.get('field', {}).get('name', '').lower()
                    if 'talle' in field_name or 'size' in field_name:
                        for value in spec.get('values', []):
                            sizes.append(value.get('name', ''))
                
                sizes = list(set(sizes))
                
                processed.append({
                    'ID': product.get('productId', ''),
                    'Nombre': product.get('productName', ''),
                    'Marca': product.get('brand', ''),
                    'Categoría': category.replace('/', ' > ').strip(' >'),
                    'Precio Lista': list_price,
                    'Precio Venta': selling_price,
                    'Descuento %': round((1 - selling_price/list_price) * 100, 0) if list_price > 0 else 0,
                    'Talles': ', '.join(sizes) if sizes else 'N/A',
                    'URL Imagen': main_image,
                    'URL Producto': f"{self.base_url}{product.get('link', '')}",
                    'Descripción': product.get('description', '')[:500]
                })
            except Exception as e:
                self.logger.warning(f"Error procesando producto {product.get('productId')}: {e}")
                continue
        
        self.logger.info(f"Productos procesados: {len(processed)}")
        return processed
    
    def create_excel(self, products: List[Dict], output_path: str, include_images: bool = False):
        """Crea archivo Excel con los productos."""
        self.logger.info(f"Creando Excel: {output_path}")
        
        wb = Workbook()
        
        # Hoja de portada
        ws_cover = wb.active
        ws_cover.title = "Portada"
        ws_cover.sheet_view.showGridLines = False
        
        # Estilos
        title_font = Font(bold=True, size=20, color="333333")
        metric_font = Font(bold=True, size=16, color="0066CC")
        
        # Título
        ws_cover.merge_cells('B2:H2')
        ws_cover['B2'] = "CATÁLOGO PORTSAID"
        ws_cover['B2'].font = title_font
        ws_cover['B2'].alignment = Alignment(horizontal='center')
        ws_cover.row_dimensions[2].height = 35
        
        # Fecha
        ws_cover.merge_cells('B4:H4')
        ws_cover['B4'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_cover['B4'].alignment = Alignment(horizontal='center')
        
        # Métricas
        ws_cover['B7'] = "Total de Productos:"
        ws_cover['D7'] = len(products)
        ws_cover['D7'].font = metric_font
        
        # Hoja de productos
        ws_data = wb.create_sheet("Productos")
        ws_data.sheet_view.showGridLines = False
        
        headers = ['ID', 'Nombre', 'Marca', 'Categoría', 'Precio Lista', 'Precio Venta', 
                   'Descuento %', 'Talles', 'URL Imagen', 'URL Producto', 'Descripción']
        
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Headers
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws_data.row_dimensions[2].height = 25
        
        # Anchos de columna
        widths = {'A': 12, 'B': 35, 'C': 15, 'D': 25, 'E': 14, 'F': 14, 'G': 12, 'H': 30, 'I': 50, 'J': 45, 'K': 45}
        for col, width in widths.items():
            ws_data.column_dimensions[col].width = width
        
        # Datos
        for row_idx, product in enumerate(products, 3):
            ws_data.cell(row=row_idx, column=1, value=product['ID']).alignment = Alignment(horizontal='center')
            ws_data.cell(row=row_idx, column=2, value=product['Nombre']).alignment = Alignment(vertical='top', wrap_text=True)
            ws_data.cell(row=row_idx, column=3, value=product['Marca']).alignment = Alignment(horizontal='center')
            ws_data.cell(row=row_idx, column=4, value=product['Categoría']).alignment = Alignment(vertical='top')
            
            cell_lista = ws_data.cell(row=row_idx, column=5, value=product['Precio Lista'])
            cell_lista.number_format = '"$"#,##0'
            
            cell_venta = ws_data.cell(row=row_idx, column=6, value=product['Precio Venta'])
            cell_venta.number_format = '"$"#,##0'
            
            cell_desc = ws_data.cell(row=row_idx, column=7, value=product['Descuento %'])
            cell_desc.number_format = '0"%"'
            
            ws_data.cell(row=row_idx, column=8, value=product['Talles']).alignment = Alignment(vertical='top', wrap_text=True)
            ws_data.cell(row=row_idx, column=9, value=product['URL Imagen']).style = 'Hyperlink'
            ws_data.cell(row=row_idx, column=10, value=product['URL Producto']).style = 'Hyperlink'
            ws_data.cell(row=row_idx, column=11, value=product['Descripción']).alignment = Alignment(vertical='top', wrap_text=True)
            
            if row_idx % 2 == 0:
                for col in range(1, 12):
                    ws_data.cell(row=row_idx, column=col).fill = alt_fill
            
            ws_data.row_dimensions[row_idx].height = 50
        
        # Data bars
        data_bar_rule = DataBarRule(start_type='min', end_type='max', color='4A90D9', showValue=True)
        ws_data.conditional_formatting.add(f'F3:F{len(products)+2}', data_bar_rule)
        
        wb.save(output_path)
        self.logger.info(f"Excel guardado: {output_path}")
    
    def create_csv_google_sheets(self, products: List[Dict], output_path: str):
        """Crea CSV con fórmulas IMAGE para Google Sheets."""
        self.logger.info(f"Creando CSV para Google Sheets: {output_path}")
        
        gs_data = []
        for product in products:
            gs_data.append({
                'Foto': f'=IMAGE("{product["URL Imagen"]}")',
                'ID': product['ID'],
                'Nombre': product['Nombre'],
                'Marca': product['Marca'],
                'Categoría': product['Categoría'],
                'Precio Lista': product['Precio Lista'],
                'Precio Venta': product['Precio Venta'],
                'Descuento %': product['Descuento %'],
                'Talles': product['Talles'],
                'URL Producto': product['URL Producto'],
                'Descripción': product['Descripción']
            })
        
        df = pd.DataFrame(gs_data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        self.logger.info(f"CSV guardado: {output_path}")
    
    def run(self, include_images: bool = False) -> Dict[str, str]:
        """Ejecuta el proceso completo de scraping."""
        start_time = time.time()
        
        try:
            # 1. Extraer productos
            raw_products = self.fetch_products()
            
            # 2. Procesar productos
            processed_products = self.process_products(raw_products)
            
            # 3. Generar timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = Path(self.config['output_dir'])
            output_dir.mkdir(exist_ok=True)
            
            # 4. Crear archivos
            outputs = {}
            
            # Excel básico
            excel_path = output_dir / f"portsaid_catalogo_{timestamp}.xlsx"
            self.create_excel(processed_products, str(excel_path), include_images)
            outputs['excel'] = str(excel_path)
            
            # CSV para Google Sheets
            csv_path = output_dir / f"portsaid_catalogo_gs_{timestamp}.csv"
            self.create_csv_google_sheets(processed_products, str(csv_path))
            outputs['csv_google_sheets'] = str(csv_path)
            
            # CSV estándar
            csv_std_path = output_dir / f"portsaid_catalogo_{timestamp}.csv"
            pd.DataFrame(processed_products).to_csv(csv_std_path, index=False, encoding='utf-8-sig')
            outputs['csv_standard'] = str(csv_std_path)
            
            elapsed = time.time() - start_time
            self.logger.info(f"Proceso completado en {elapsed:.1f} segundos")
            self.logger.info(f"Archivos generados: {outputs}")
            
            return outputs
            
        except Exception as e:
            self.logger.error(f"Error en el proceso: {e}", exc_info=True)
            raise


def main():
    """Función principal."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Portsaid Catalog Scraper')
    parser.add_argument('--config', '-c', help='Ruta al archivo de configuración JSON')
    parser.add_argument('--images', '-i', action='store_true', help='Incluir imágenes en el Excel')
    parser.add_argument('--output', '-o', default='output', help='Directorio de salida')
    
    args = parser.parse_args()
    
    # Crear scraper
    scraper = PortsaidScraper(config_path=args.config)
    scraper.config['output_dir'] = args.output
    
    # Ejecutar
    outputs = scraper.run(include_images=args.images)
    
    print("\n" + "="*50)
    print("ARCHIVOS GENERADOS:")
    print("="*50)
    for key, path in outputs.items():
        print(f"  {key}: {path}")


if __name__ == "__main__":
    main()
